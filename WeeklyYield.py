import time
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import mysql.connector

mydb=mysql.connector.connect(
    host="localhost",
    user="root",
    password="19071234",
    database="uygulama1"
)
mycursor=mydb.cursor()

def girispaneli():
    o=input("Giriş yapmak için 1'e,Kayıt olmak için ise 2'ye tıklayın: ")
    if o=="1":
        
    if o=="2":
        adi=input("Lütfen Adınızı Girin: ")
        sifre=input("Lütfen Şifrenizi Girin")
        sql="INSERT INTO uygulama1 (ad,sifre) VALUES (%s,%s)"
        val=(adi,sifre)
        mycursor.execute(sql,val)
        mydb.commit()
        print(mycursor.rowcount," Kaydınız başarıyla eklenmiştir")
    
    else:
        girispaneli()
    



def anauygulama():
    print("Uygulamamıza Hoşgeldiniz!")

    girispaneli()

    giris2=input("Kronometre Uygulaması için A'ya tıklayın,Haftalık Verim Hesaplayıcı için ise B'ye tıklayın -->")
    if giris2=="A":
        KronometreUygulamasi()
    elif giris2=="a":
        KronometreUygulamasi()
    elif giris2=="B":
        HaftalıkVerimHesaplama()
    elif giris2=="b":
        HaftalıkVerimHesaplama()
    else:
        anauygulama()






def KronometreUygulamasi():
    a = " "
    gir = input("Kronometreyi başlatmak için T'ye tıklayın -->")
    h=gir.lower()
    if h == "t":
        bas=time.strftime('%X')
        print("Başlangıç Zamanı -->", bas)
        bel= input("Kronometreyi durdurmak için X'e tıklayın ->")
        if bel == "X":
            son = time.strftime('%X')
            print("Bitiş Zamanı -->", son)
            bassaat = int(bas[0:2])
            basdk = int(bas[3:5])
            sonsaat = int(son[0:2])
            sondk = int(son[3:5])
            saat = sonsaat - bassaat
            dakika = sondk - basdk
            t=time.strftime('%x') + " Toplam Çalışılan Süre " + str(saat) + "saat" + "-" + str(dakika) + "dakika"
            print(str(t))
            print("\n")

            A=time.strftime('%x')
            B=A[0:2]
            C=A[3:5]
            D=A[6:8]
            dosya2 = open(str(B+" "+C+" "+D)+".txt","w")
            dosya2.write(str(t))

            KronometreUygulamasi()
    else:
        pass




def HaftalıkVerimHesaplama():
    print("Günlerin önüne kaç saat çalıştığınızı yazın")

    try:
        a=float(input("Pazartesi -> "))
        b=float(input("Salı -> "))
        c=float(input("Çarşamba -> "))         
        d=float(input("Perşembe -> "))
        e=float(input("Cuma -> "))
        f=float(input("Cumartesi -> "))
        g=float(input("Pazar -> "))

    except ValueError:
        print("Lütfen Bir Sayı Giriniz")

    gunlukort = float(a + b + c + d + e + f + g) / 7
    hours=[a,b,c,d,e,f,g]
    gunler=["Pazartesi","Salı","Çarşamba","Perşembe","Cuma","Cumartesi","Pazar"]       
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(7):
        c1=sheet.cell(row=i+1,column=1)
        c1.value=gunler[i]

        c2=sheet.cell(row=i+1,column=2)
        c2.value = hours[i]

        c3=sheet.cell(row=i+1,column=3)
        c3.value=hours[i]*60

    c4=sheet.cell(row=7,column=1)
    c4.value="Haftalık rapor"

    c5=sheet.cell(row=7,column=2)
    c5.value=gunlukort

    wb.save("rapor.xlsx")



    for i in range(len(hours)-1):
        if hours[i]>23:
            print("Hatalı Veri Girişi")
            HaftalıkVerimHesaplama()
        else:
            pass


    if gunlukort>=24:
        print("HATALI VERİ GİRMİŞ OLABİLİRSİNİZ.LÜTFEN GİRDİĞİNİZ VERİLERİ KONTROL EDİN")
        HaftalıkVerimHesaplama()
    else:
        saat = int(gunlukort)
        dakika = gunlukort - int(gunlukort)
        dk = (saat * 60) + dakika
        print("Bu hafta Günlük Ortalama ", int(dk), " dakika çalışıldı","\n")


    giris=input("Haftalık Verim Grafiğini Görmek için Q'ya,Haftalık Raporu Görmek için ise E'ye tıklayın --->")
    if giris=="Q":
        haftx = np.array(["PTESİ","SALI","ÇŞAMBA","PŞEMBE","CUMA","CTESİ","PAZAR"])
        saaty = np.array([a, b, c, d, e, f, g])
        plt.plot(haftx, saaty)
        plt.show()
    elif giris=="E":
        dosya1 = load_workbook("rapor.xlsx")
        sheet = dosya1.active

        dosya = load_workbook("rapor.xlsx")
        sheet = dosya.active
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=7, max_col=2):
            for cell in row:
                print(cell.value, end=" ")
            print()

            dosya.close()

    else:
        pass



anauygulama()


